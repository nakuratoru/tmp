# ====== 設定 ======
PADDING_MODE = "same"  # "same (with padding)" のみ
EXCEL_INPUT_FILE = "cnn.xlsx"
EXCEL_OUTPUT_FILE = "cnn_filters_simple4.xlsx"
SHEET_NAMES = {
    "input": "original",
    "vertical": "vertical",
    "horizontal": "horizontal",
    "diag45": "diag45",
    "diag135": "diag135",
}
# ===================

import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill

# Excel から入力シート読み込み
orig_df = pd.read_excel(EXCEL_INPUT_FILE, sheet_name=SHEET_NAMES["input"], header=None)
orig = orig_df.fillna(0).to_numpy()

def xcorr2d_same(img, kernel):
    return xcorr2d(img, kernel, padding="same")

def xcorr2d(img, kernel, padding="same"):
    kh, kw = kernel.shape
    H, W = img.shape
    if padding == "same":
        pad_h = kh // 2
        pad_w = kw // 2
        padded = np.pad(img, ((pad_h, kh-1-pad_h), (pad_w, kw-1-pad_w)), mode='constant')
        out_h, out_w = H, W
    else:
        raise ValueError("padding must be 'same'")  # validはサポートしない
    out = np.zeros((out_h, out_w), dtype=float)
    for i in range(out_h):
        for j in range(out_w):
            region = padded[i:i+kh, j:j+kw]
            out[i, j] = np.sum(region * kernel)
    return out

def pad_bottom_right(arr):
    """配列の下と右に0を1行ずつ追加"""
    h, w = arr.shape
    out = np.zeros((h+1, w+1), dtype=arr.dtype)
    out[:h, :w] = arr
    return out

# 4つのシンプルフィルタ
K_vert   = np.array([[0,0,0],[0,-1,1],[0,0,0]])
K_horiz  = np.array([[0,0,0],[0,-1,0],[0,1,0]])
K_diag45 = np.array([[0,0,0],[0,-1,0],[0,0,1]])
K_diag135= np.array([[0,0,0],[0,-1,0],[1,0,0]])

F_vert   = xcorr2d(orig, K_vert, padding=PADDING_MODE)
F_horiz  = xcorr2d(orig, K_horiz, padding=PADDING_MODE)
F_diag45 = xcorr2d(orig, K_diag45, padding=PADDING_MODE)
F_diag135= xcorr2d(orig, K_diag135, padding=PADDING_MODE)

F_vert_padded   = pad_bottom_right(F_vert)
F_horiz_padded  = pad_bottom_right(F_horiz)
F_diag45_padded = pad_bottom_right(F_diag45)
F_diag135_padded= pad_bottom_right(F_diag135)

# Excel 出力
with pd.ExcelWriter(EXCEL_OUTPUT_FILE, engine="openpyxl") as writer:
    pd.DataFrame(orig).to_excel(writer, sheet_name=SHEET_NAMES["input"], index=False, header=False)
    pd.DataFrame(F_vert_padded).to_excel(writer, sheet_name=SHEET_NAMES["vertical"], index=False, header=False)
    pd.DataFrame(F_horiz_padded).to_excel(writer, sheet_name=SHEET_NAMES["horizontal"], index=False, header=False)
    pd.DataFrame(F_diag45_padded).to_excel(writer, sheet_name=SHEET_NAMES["diag45"], index=False, header=False)
    pd.DataFrame(F_diag135_padded).to_excel(writer, sheet_name=SHEET_NAMES["diag135"], index=False, header=False)

    # 値が10は薄い青、-10は薄い赤で塗る
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 各シートに対してセルを走査して色付け
    sheets = {
        SHEET_NAMES["input"]: orig,
        SHEET_NAMES["vertical"]: F_vert_padded,
        SHEET_NAMES["horizontal"]: F_horiz_padded,
        SHEET_NAMES["diag45"]: F_diag45_padded,
        SHEET_NAMES["diag135"]: F_diag135_padded,
    }

    for name, arr in sheets.items():
        ws = writer.sheets[name]
        rows, cols = arr.shape
        # 列幅を1.7に設定
        for c in range(1, cols + 1):
            col_letter = ws.cell(row=1, column=c).column_letter
            ws.column_dimensions[col_letter].width = 2.4
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v == 10 or v == 10.0:
                    cell.fill = blue_fill
                elif v == -10 or v == -10.0:
                    cell.fill = red_fill
